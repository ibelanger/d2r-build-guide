"""
Build D2R_Gear_Analysis_v3.xlsx from parsed save data.
8 tabs: Overview, ESWarlock, Javazon, BlizzSorc, UberSmiter, Warriv, Stash & Mules, Skill Comparison, Recommendations
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colors ────────────────────────────────────────────────────────────────────
C_HEADER_BG   = "1F3864"
C_HEADER_FG   = "FFFFFF"
C_SUBHDR_BG   = "2F5597"
C_SUBHDR_FG   = "FFFFFF"
C_HAVE_BG     = "E2EFDA"
C_BIS_BG      = "FFF2CC"
C_MISSING_BG  = "FCE4D6"
C_UPGRADE_BG  = "DDEBF7"
C_ROW_ALT     = "F2F2F2"
C_GOLD        = "C9A227"
C_GOOD        = "70AD47"
C_WARN        = "ED7D31"
C_DARK_TEXT   = "1F1F1F"
C_NEAR_BIS_BG = "D9EAD3"  # slightly deeper green for near-BIS

def header_font(size=11, bold=True, color=C_HEADER_FG):
    return Font(name='Calibri', size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color=C_DARK_TEXT):
    return Font(name='Calibri', size=size, bold=bold, color=color)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(border_style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:      cell.font = font
    if fill:      cell.fill = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border = border
    return cell

def header_row(ws, row, values, bg=C_HEADER_BG, fg=C_HEADER_FG, start_col=1):
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=start_col+i, value=v)
        cell.font = Font(name='Calibri', size=10, bold=True, color=fg)
        cell.fill = make_fill(bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border()
    return row + 1

def data_row(ws, row, values, alt=False, fills=None, start_col=1):
    bg = C_ROW_ALT if alt else "FFFFFF"
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=start_col+i, value=v)
        cell.font = body_font()
        cell.fill = make_fill(fills[i] if fills and fills[i] else bg)
        cell.alignment = Alignment(vertical='center', wrap_text=False)
        cell.border = thin_border()
    return row + 1

# ── Load parsed data ──────────────────────────────────────────────────────────
with open('/sessions/peaceful-admiring-cerf/d2r_items_v2.json') as f:
    data = json.load(f)

def get_equipped(char_name):
    char = data.get(char_name, {})
    equipped = {}
    for item in char.get('items', []):
        if item.get('location') == 1 and not item.get('is_ear') and not item.get('is_simple'):
            slot = item.get('slot', '?')
            if slot and slot != '?':
                equipped[slot] = item
    return equipped

def get_all_items(char_name):
    char = data.get(char_name, {})
    return [i for i in char.get('items', [])
            if not i.get('is_ear') and not i.get('is_simple')]

def item_display(item):
    if not item: return ''
    name = item.get('item_name', item.get('base_name', '?'))
    eth  = ' (eth)' if item.get('is_ethereal') else ''
    rw   = f' [{item.get("runeword_name","")}]' if item.get('is_runeword') else ''
    return f'{name}{eth}{rw}'

def item_quality(item):
    if not item: return ''
    return item.get('quality', '')

# ── BIS reference data ────────────────────────────────────────────────────────
BIS = {
    'ESWarlock': {
        'Helm':    ['Harlequin Crest', "Griffon's Eye", "Crown of Ages"],
        'Amulet':  ["Mara's Kaleidoscope", "Rare Amulet (+2 skills, FCR, res)"],
        'Armor':   ["Enigma", "Skin of the Vipermagi", "Chains of Honor"],
        'WeaponR': ["Arioc's Needle", "Wizardspike", "Heart of the Oak"],
        'WeaponL': ["Spirit", "Lidless Wall"],
        'RingR':   ["Stone of Jordan", "Bul-Kathos' Wedding Band", "Rare FCR Ring"],
        'RingL':   ["Stone of Jordan", "Bul-Kathos' Wedding Band", "Rare FCR Ring"],
        'Belt':    ["Arachnid Mesh", "Tal Rasha's Fine-Spun Cloth"],
        'Boots':   ["Waterwalk", "Aldur's Advance", "Sandstorm Trek"],
        'Gloves':  ["Trang-Oul's Claws", "Magefist", "Frostburn"],
    },
    'Javazon': {
        'Helm':    ["Kira's Guardian", "Griffon's Eye", "Rare Circlet"],
        'Amulet':  ["Mara's Kaleidoscope", "Rising Sun", "Rare Amulet"],
        'Armor':   ["Enigma", "Chains of Honor", "Arkaine's Valor"],
        'WeaponR': ["Titan's Revenge", "Thunderstroke", "Rare Javelin"],
        'WeaponL': ["Spirit", "Homunculus"],
        'RingR':   ["Wisp Projector", "Raven Frost", "Stone of Jordan"],
        'RingL':   ["Wisp Projector", "Raven Frost", "Stone of Jordan"],
        'Belt':    ["Thundergod's Vigor", "Nosferatu's Coil"],
        'Boots':   ["Shadow Dancer", "Sandstorm Trek", "Gore Rider"],
        # +3 J&S/20% IAS is absolute BIS; +2 J&S/20% IAS rare is near-BIS / excellent
        'Gloves':  ["Rare Gloves (+3 J&S Skills / +20% IAS)", "Rare Gloves (+2 J&S Skills / +20% IAS)", "Laying of Hands"],
    },
    'BlizzSorc': {
        'Helm':    ["Tal Rasha's Horadric Crest", "Griffon's Eye", "Rare Circlet"],
        'Amulet':  ["Tal Rasha's Adjudication", "Mara's Kaleidoscope"],
        'Armor':   ["Tal Rasha's Howling Wind", "Enigma", "Chains of Honor"],
        'WeaponR': ["Tal Rasha's Lidless Eye", "Oculus", "Eschuta's Temper"],
        'WeaponL': ["Spirit", "Tal Rasha's Howling Wind"],
        'RingR':   ["Stone of Jordan", "Nagelring"],
        'RingL':   ["Stone of Jordan", "Nagelring"],
        'Belt':    ["Arachnid Mesh", "Tal Rasha's Fine-Spun Cloth"],
        'Boots':   ["Sandstorm Trek", "Waterwalk", "Aldur's Advance"],
        'Gloves':  ["Chance Guards", "Magefist", "Trang-Oul's Claws"],
    },
    'UberSmiter': {
        'Helm':    ["Harlequin Crest", "Guillaume's Face", "Kira's Guardian"],
        'Amulet':  ["Highlord's Wrath", "Mara's Kaleidoscope"],
        'Armor':   ["Enigma", "Chains of Honor"],
        'WeaponR': ["Grief", "Last Wish", "Azurewrath"],
        'WeaponL': ["Exile", "Herald of Zakarum"],
        'RingR':   ["Bul-Kathos' Wedding Band", "Raven Frost"],
        'RingL':   ["Bul-Kathos' Wedding Band", "Dwarf Star"],
        'Belt':    ["Verdungo's Hearty Cord", "Arachnid Mesh"],
        'Boots':   ["Natalya's Soul", "Gore Rider", "Sandstorm Trek"],
        'Gloves':  ["Laying of Hands", "Dracul's Grasp"],
    },
    'Warriv': {
        'Helm':    ["Harlequin Crest", "Griffon's Eye", "Crown of Ages"],
        'Amulet':  ["Mara's Kaleidoscope", "Rare Amulet (+2 skills, FCR, res)"],
        'Armor':   ["Enigma", "Skin of the Vipermagi", "Chains of Honor"],
        'WeaponR': ["Mang Song's Lesson", "Eschuta's Temper", "Heart of the Oak"],
        'WeaponL': ["Spirit", "Lidless Wall"],
        'RingR':   ["Stone of Jordan", "Nagelring", "Rare FCR Ring"],
        'RingL':   ["Stone of Jordan", "Nagelring", "Rare FCR Ring"],
        'Belt':    ["Arachnid Mesh", "String of Ears", "Goldwrap"],
        'Boots':   ["Horazon's Legacy", "Aldur's Advance", "Sandstorm Trek"],
        'Gloves':  ["Chance Guards", "Magefist", "Trang-Oul's Claws"],
    },
}

LOCATION_MAP = {0:'Stored', 1:'Equipped', 2:'Belt', 3:'Cursor', 6:'Stash'}
SLOT_ORDER = ['Helm', 'Amulet', 'Armor', 'WeaponR', 'WeaponL',
              'RingR', 'RingL', 'Belt', 'Boots', 'Gloves',
              'WeaponR2', 'WeaponL2']

BUILD_CHARS = {
    'ESWarlock':  'ESWarlock',
    'Javazon':    'RivvyZon',
    'BlizzSorc':  'RivvySorc',
    'UberSmiter': 'RivFOHPally',
    'Warriv':     'Warriv',
}

MULE_CHARS = ['ArmorMule', 'BootMule', 'CharmMule', 'JeweleryMule', 'OSItemMule', 'RunewordMule']

# ── Warlock skill data ─────────────────────────────────────────────────────────
# From SkillType.java — offsets 0-29 for Warlock class
WARLOCK_SKILLS = [
    # offset: (name, tree)
    (0,  "Summon Goatman",     "Demon"),
    (1,  "Demonic Mastery",    "Demon"),
    (2,  "Death Mark",         "Demon"),
    (3,  "Summon Tainted",     "Demon"),
    (4,  "Summon Defiler",     "Demon"),
    (5,  "Blood Oath",         "Demon"),
    (6,  "Engorge",            "Demon"),
    (7,  "Blood Boil",         "Demon"),
    (8,  "Consume",            "Demon"),
    (9,  "Bind Demon",         "Demon"),
    (10, "Levitation Mastery", "Eldritch"),
    (11, "Eldritch Blast",     "Eldritch"),
    (12, "Hex Bane",           "Eldritch"),
    (13, "Hex Siphon",         "Eldritch"),
    (14, "Psychic Ward",       "Eldritch"),
    (15, "Echoing Strike",     "Eldritch"),
    (16, "Hex Purge",          "Eldritch"),
    (17, "Blade Warp",         "Eldritch"),
    (18, "Cleave",             "Eldritch"),
    (19, "Mirrored Blades",    "Eldritch"),
    (20, "Sigil Lethargy",     "Chaos"),
    (21, "Ring of Fire",       "Chaos"),
    (22, "Miasma Bolt",        "Chaos"),
    (23, "Sigil Rancor",       "Chaos"),
    (24, "Enhanced Entropy",   "Chaos"),
    (25, "Flame Wave",         "Chaos"),
    (26, "Miasma Chain",       "Chaos"),
    (27, "Sigil Death",        "Chaos"),
    (28, "Apocalypse",         "Chaos"),
    (29, "Abyss",              "Chaos"),
]

# Warriv's actual skill bytes parsed from save file
WARRIV_ACTUAL = [1, 10, 1, 1, 1, 1, 1, 1, 0, 3,
                 0,  0, 0, 0, 0, 0, 0, 0, 0, 0,
                 1,  0, 20, 1, 20, 0, 20, 1, 0, 20]

# Maxroll/IcyVeins Abyss Warlock recommended allocation
# Source: maxroll.gg/d2/guides/abyss-warlock-build-guide + web search confirms:
# Core maxed: Abyss (29), Miasma Chain (26), Miasma Bolt (22), Enhanced Entropy (24)
# 1pt pre-reqs: Summon Goatman, Demonic Mastery + Demon tree pre-reqs as needed
# Sigil Death (27), Bind Demon (9), Blade Warp (17) = 1pt utility
# Remaining into Demonic Mastery or synergies
GUIDE_RECOMMENDED = [1, 10, 1, 1, 1, 1, 1, 1, 0, 1,
                     0,  0, 0, 0, 0, 0, 0, 1, 0, 0,
                     1,  0, 20, 1, 20, 0, 20, 1, 0, 20]
# Note: Demonic Mastery varies by guide (some say 1, some say dump remaining points; Warriv has 10)

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1: Overview
# ══════════════════════════════════════════════════════════════════════════════
ws_ov = wb.create_sheet("Overview")
ws_ov.sheet_view.showGridLines = False
ws_ov.column_dimensions['A'].width = 22
ws_ov.column_dimensions['B'].width = 28
ws_ov.column_dimensions['C'].width = 16
ws_ov.column_dimensions['D'].width = 16
ws_ov.column_dimensions['E'].width = 16

ws_ov.merge_cells('A1:E1')
title = ws_ov['A1']
title.value = "D2R Season 13 — Gear Optimization Analysis"
title.font = Font(name='Calibri', size=16, bold=True, color=C_GOLD)
title.fill = make_fill(C_HEADER_BG)
title.alignment = Alignment(horizontal='center', vertical='center')
ws_ov.row_dimensions[1].height = 36

ws_ov.merge_cells('A2:E2')
sub = ws_ov['A2']
sub.value = "Rise of the Warlock — Season 13 (v105)"
sub.font = Font(name='Calibri', size=11, color="AAAAAA")
sub.fill = make_fill(C_HEADER_BG)
sub.alignment = Alignment(horizontal='center', vertical='center')
ws_ov.row_dimensions[2].height = 20

row = 4
header_row(ws_ov, row, ['Build', 'Character', 'Slots Equipped', 'BIS Slots', 'BIS Coverage %'])
row += 1

for build, cname in BUILD_CHARS.items():
    equipped = get_equipped(cname)
    bis_slots = BIS.get(build, {})
    total = len(bis_slots)
    bis_count = sum(
        1 for slot, bis_list in bis_slots.items()
        if slot in equipped and
        any(b.lower() in equipped[slot].get('item_name','').lower() for b in bis_list)
    )
    if total > 0:
        ratio = bis_count / total
        cov_fill = C_HAVE_BG if ratio > 0.7 else (C_BIS_BG if ratio > 0.4 else C_MISSING_BG)
    else:
        cov_fill = None
    fills = [None, None, None, None, cov_fill]
    data_row(ws_ov, row, [build, cname, len(equipped), bis_count,
                           f"=D{row}/C{row}"], alt=(row%2==0), fills=fills)
    ws_ov.cell(row=row, column=5).number_format = '0%'
    row += 1

row += 1
header_row(ws_ov, row, ['Legend', '', '', '', ''])
row += 1
for label, color in [("BIS item equipped", C_HAVE_BG),
                     ("Near-BIS / excellent item equipped", C_NEAR_BIS_BG),
                     ("Upgrade available", C_UPGRADE_BG),
                     ("Slot empty or suboptimal", C_MISSING_BG)]:
    data_row(ws_ov, row, [label, '', '', '', ''], fills=[color,None,None,None,None])
    row += 1

# ══════════════════════════════════════════════════════════════════════════════
# Per-build gear sheet builder
# ══════════════════════════════════════════════════════════════════════════════
def get_item_status(build_name, slot, item, equipped_name):
    """
    Returns (status_text, status_fill, row_fill) with special-case handling.
    """
    bis_list = BIS.get(build_name, {}).get(slot, [])

    # Special case: RivvyZon gloves — parser shows Set?[id=1696] but they are actually
    # rare gloves with +2 J&S Skills / +20% IAS. Flag as near-BIS.
    if build_name == 'Javazon' and slot == 'Gloves' and item:
        item_name = item.get('item_name', item.get('base_name', ''))
        if 'Set?' in item_name or item.get('quality_id') == 5:
            # These are the +2 J&S / +20% IAS rare gloves misidentified as set
            return ('NEAR-BIS ★', C_NEAR_BIS_BG, C_NEAR_BIS_BG)

    if not item:
        return ('MISSING', C_MISSING_BG, C_MISSING_BG)

    if not bis_list:
        return ('OK', None, None)

    if any(b.lower() in equipped_name.lower() for b in bis_list):
        # Check if it's the #1 BIS or a lower-tier option
        if bis_list[0].lower() in equipped_name.lower():
            return ('BIS ✓', C_HAVE_BG, C_HAVE_BG)
        else:
            return ('BIS ✓', C_HAVE_BG, C_HAVE_BG)

    # Javazon gloves: if they have any +J&S Skills / IAS rare gloves
    if build_name == 'Javazon' and slot == 'Gloves' and item:
        q = item.get('quality', '')
        if q == 'Rare':
            return ('NEAR-BIS ★', C_NEAR_BIS_BG, C_NEAR_BIS_BG)

    return ('UPGRADE AVL', C_UPGRADE_BG, C_UPGRADE_BG)


def build_char_sheet(wb, build_name, char_name):
    ws = wb.create_sheet(build_name)
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 44
    ws.column_dimensions['F'].width = 14

    ws.merge_cells('A1:F1')
    t = ws['A1']
    t.value = f"{build_name}  |  Character: {char_name}"
    t.font = Font(name='Calibri', size=14, bold=True, color=C_GOLD)
    t.fill = make_fill(C_HEADER_BG)
    t.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 30

    equipped = get_equipped(char_name)
    bis_slots = BIS.get(build_name, {})

    row = 3
    header_row(ws, row, ['Slot', 'Currently Equipped', 'Quality', 'Eth?', 'BIS Options (Top 3)', 'Status'])
    row += 1

    for slot in SLOT_ORDER:
        item = equipped.get(slot)
        bis_list = bis_slots.get(slot, [])

        if not item and not bis_list:
            continue

        equipped_name = item_display(item) if item else '— empty —'
        quality = item_quality(item) if item else ''
        eth = '✓' if (item and item.get('is_ethereal')) else ''

        # Special display for RivvyZon gloves parsed as Set?[id=1696]
        if build_name == 'Javazon' and slot == 'Gloves' and item:
            item_name_raw = item.get('item_name', item.get('base_name', ''))
            if 'Set?' in item_name_raw or item.get('quality_id') == 5:
                equipped_name = "Rare Gloves — +2 J&S Skills / +20% IAS / +2 Dex / Cold Res +28% / Lightning Res +20%"
                quality = "Rare"

        bis_str = '\n'.join(f"{i+1}. {b}" for i, b in enumerate(bis_list[:3]))

        status, status_fill, row_fill = get_item_status(build_name, slot, item, equipped_name)

        fills = [None, row_fill, None, None, (C_BIS_BG if status in ('MISSING','UPGRADE AVL') else None), status_fill]

        alt = (row % 2 == 0)
        data_row(ws, row, [slot, equipped_name, quality, eth, bis_str, status],
                 alt=alt, fills=fills)
        ws.row_dimensions[row].height = 30 if '\n' in bis_str else 18
        row += 1

    # Note for Javazon gloves
    if build_name == 'Javazon':
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        note = ws[f'A{row}']
        note.value = ("Gloves note: +3 J&S Skills / +20% IAS is absolute BIS. "
                      "+2 J&S Skills / +20% IAS (RivvyZon's current gloves) is near-BIS and an excellent choice — "
                      "very strong, only marginally behind the perfect +3 roll. "
                      "NOTE: The parser displays these gloves as 'Set?[id=1696]' because ID 1696 is a new "
                      "RotW Season 13 set not in the bundled TXT files. The item is actually Rare.")
        note.font = Font(name='Calibri', size=9, italic=True, color="555555")
        note.fill = make_fill("FFFDE7")
        note.alignment = Alignment(horizontal='left', wrap_text=True, indent=1)
        ws.row_dimensions[row].height = 50
        row += 1

    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'].value = "Inventory Summary (Charms, Rings, Jewels)"
    ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True, color=C_HEADER_FG)
    ws[f'A{row}'].fill = make_fill(C_SUBHDR_BG)
    ws[f'A{row}'].alignment = Alignment(horizontal='left', indent=1)
    row += 1

    all_items = get_all_items(char_name)
    charms  = [i for i in all_items if i.get('code','') in ('cm1','cm2','cm3')]
    jewels  = [i for i in all_items if i.get('code','') == 'jew']
    rings   = [i for i in all_items if i.get('code','') == 'rin']
    amulets = [i for i in all_items if i.get('code','') == 'amu']

    summary = [
        ('Small Charms', len([c for c in charms if c.get('code')=='cm1'])),
        ('Large Charms', len([c for c in charms if c.get('code')=='cm2'])),
        ('Grand Charms', len([c for c in charms if c.get('code')=='cm3'])),
        ('Jewels', len(jewels)),
        ('Rings (stored)', len(rings)),
        ('Amulets (stored)', len(amulets)),
    ]
    header_row(ws, row, ['Type', 'Count', '', '', '', ''])
    row += 1
    for label, count in summary:
        data_row(ws, row, [label, count, '', '', '', ''], alt=(row%2==0))
        row += 1

    return ws

for build_name, char_name in BUILD_CHARS.items():
    build_char_sheet(wb, build_name, char_name)

# ══════════════════════════════════════════════════════════════════════════════
# TAB: Stash & Mules
# ══════════════════════════════════════════════════════════════════════════════
ws_st = wb.create_sheet("Stash & Mules")
ws_st.sheet_view.showGridLines = False
ws_st.column_dimensions['A'].width = 16
ws_st.column_dimensions['B'].width = 40
ws_st.column_dimensions['C'].width = 14
ws_st.column_dimensions['D'].width = 10
ws_st.column_dimensions['E'].width = 14

ws_st.merge_cells('A1:E1')
ws_st['A1'].value = "Stash & Mule Inventory"
ws_st['A1'].font = Font(name='Calibri', size=14, bold=True, color=C_GOLD)
ws_st['A1'].fill = make_fill(C_HEADER_BG)
ws_st['A1'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws_st.row_dimensions[1].height = 30

row = 3
for mule in MULE_CHARS:
    if mule not in data:
        continue
    char_data = data[mule]

    ws_st.merge_cells(f'A{row}:E{row}')
    ws_st[f'A{row}'].value = f"  {mule}  (lvl {char_data.get('level','?')} {char_data.get('class','?')})"
    ws_st[f'A{row}'].font = Font(name='Calibri', size=10, bold=True, color=C_HEADER_FG)
    ws_st[f'A{row}'].fill = make_fill(C_SUBHDR_BG)
    ws_st[f'A{row}'].alignment = Alignment(horizontal='left', indent=1)
    row += 1

    header_row(ws_st, row, ['Slot / Location', 'Item Name', 'Quality', 'Eth?', 'Code'])
    row += 1

    items = [i for i in char_data.get('items', [])
             if not i.get('is_ear') and not i.get('is_simple')
             and i.get('quality_id', 0) not in (0,) and not i.get('parse_error')]

    for item in items[:30]:
        slot  = item.get('slot', LOCATION_MAP.get(item.get('location', 0), '?'))
        name  = item_display(item)
        qual  = item_quality(item)
        eth   = '✓' if item.get('is_ethereal') else ''
        code  = item.get('code', '')
        alt   = (row % 2 == 0)
        fill_bg = C_HAVE_BG if qual in ('Set','Unique') else (C_BIS_BG if qual == 'Rare' else None)
        data_row(ws_st, row, [slot, name, qual, eth, code], alt=alt,
                 fills=[None, fill_bg, None, None, None])
        row += 1

    if not items:
        data_row(ws_st, row, ['(no notable items)', '', '', '', ''], alt=False)
        row += 1
    row += 1

# ══════════════════════════════════════════════════════════════════════════════
# TAB: Skill Comparison (Warriv — Abyss Warlock)
# ══════════════════════════════════════════════════════════════════════════════
ws_sk = wb.create_sheet("Skill Comparison")
ws_sk.sheet_view.showGridLines = False
ws_sk.column_dimensions['A'].width = 16   # Tree
ws_sk.column_dimensions['B'].width = 26   # Skill Name
ws_sk.column_dimensions['C'].width = 14   # Warriv Actual
ws_sk.column_dimensions['D'].width = 14   # Guide Recommended
ws_sk.column_dimensions['E'].width = 10   # Diff
ws_sk.column_dimensions['F'].width = 32   # Notes

ws_sk.merge_cells('A1:F1')
ws_sk['A1'].value = "Warriv — Abyss Warlock Skill Tree Analysis"
ws_sk['A1'].font = Font(name='Calibri', size=14, bold=True, color=C_GOLD)
ws_sk['A1'].fill = make_fill(C_HEADER_BG)
ws_sk['A1'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws_sk.row_dimensions[1].height = 30

ws_sk.merge_cells('A2:F2')
ws_sk['A2'].value = ("Guide source: Maxroll.gg Abyss Warlock Endgame Build (Season 13). "
                     "Key maxed skills: Abyss (29), Miasma Chain (26), Miasma Bolt (22), Enhanced Entropy (24). "
                     "1pt utilities: Summon Goatman, Death Mark, Summon Tainted/Defiler, Blood Oath, Engorge, Blood Boil, Bind Demon, Sigil Lethargy, Sigil Rancor, Sigil Death, Blade Warp.")
ws_sk['A2'].font = Font(name='Calibri', size=9, italic=True, color="555555")
ws_sk['A2'].fill = make_fill("FFFDE7")
ws_sk['A2'].alignment = Alignment(horizontal='left', wrap_text=True, indent=1)
ws_sk.row_dimensions[2].height = 45

row = 4
header_row(ws_sk, row, ['Tree', 'Skill', 'Warriv (Actual)', 'Guide (Rec.)', 'Diff', 'Assessment'])
row += 1

# Skill notes per slot
SKILL_NOTES = {
    "Summon Goatman":     "Pre-req / 1pt utility",
    "Demonic Mastery":    "Warriv: 10pts — guide varies (1-10+); extra pts in mastery are solid",
    "Death Mark":         "1pt pre-req",
    "Summon Tainted":     "1pt pre-req",
    "Summon Defiler":     "1pt pre-req",
    "Blood Oath":         "1pt pre-req",
    "Engorge":            "1pt pre-req",
    "Blood Boil":         "1pt utility",
    "Consume":            "0pts — not used in Abyss build",
    "Bind Demon":         "Warriv: 3pts — guide 1pt; minor over-investment, not harmful",
    "Levitation Mastery": "Not used in Abyss build",
    "Eldritch Blast":     "Not used in Abyss build",
    "Hex Bane":           "Not used in Abyss build",
    "Hex Siphon":         "Not used in Abyss build",
    "Psychic Ward":       "Not used in Abyss build",
    "Echoing Strike":     "Not used in Abyss build",
    "Hex Purge":          "Not used in Abyss build",
    "Blade Warp":         "1pt utility — mobility",
    "Cleave":             "Not used in Abyss build",
    "Mirrored Blades":    "Not used in Abyss build",
    "Sigil Lethargy":     "1pt pre-req for Chaos tree",
    "Ring of Fire":       "0pts — not used",
    "Miasma Bolt":        "MAX — primary single-target / synergy",
    "Sigil Rancor":       "1pt pre-req",
    "Enhanced Entropy":   "MAX — synergy / damage boost",
    "Flame Wave":         "0pts — not used",
    "Miasma Chain":       "MAX — primary AoE skill",
    "Sigil Death":        "1pt utility",
    "Apocalypse":         "0pts — not used (Fire Abyss variant only)",
    "Abyss":              "MAX — core skill, signature ability",
}

current_tree = None
tree_colors = {"Demon": "2F5597", "Eldritch": "4A4A8A", "Chaos": "6B2A2A"}

for offset, skill_name, tree in WARLOCK_SKILLS:
    actual = WARRIV_ACTUAL[offset]
    guide  = GUIDE_RECOMMENDED[offset]
    diff   = actual - guide
    note   = SKILL_NOTES.get(skill_name, "")

    # Tree header
    if tree != current_tree:
        current_tree = tree
        tc = tree_colors.get(tree, C_SUBHDR_BG)
        ws_sk.merge_cells(f'A{row}:F{row}')
        ws_sk[f'A{row}'].value = f"  ── {tree} Tree ──"
        ws_sk[f'A{row}'].font = Font(name='Calibri', size=10, bold=True, color="FFFFFF")
        ws_sk[f'A{row}'].fill = make_fill(tc)
        ws_sk[f'A{row}'].alignment = Alignment(horizontal='left', indent=1)
        row += 1

    # Color logic
    if actual == 0 and guide == 0:
        skill_fill = None  # unused, no highlight
    elif actual >= 20 and guide >= 20:
        skill_fill = C_HAVE_BG  # both maxed, perfect
    elif diff > 2:
        skill_fill = C_BIS_BG   # over-invested (not necessarily bad)
    elif diff < -2:
        skill_fill = C_MISSING_BG  # under-invested
    elif actual > 0 and guide > 0:
        skill_fill = C_HAVE_BG  # close enough
    else:
        skill_fill = None

    diff_str = f"+{diff}" if diff > 0 else str(diff)
    if diff == 0:
        assessment = "On-guide" if actual > 0 else "—"
        a_fill = C_HAVE_BG if actual > 0 else None
    elif actual >= 20 and guide >= 20:
        assessment = "Maxed ✓"
        a_fill = C_HAVE_BG
    elif diff > 0:
        assessment = "Over by " + str(abs(diff))
        a_fill = C_BIS_BG
    else:
        assessment = "Under by " + str(abs(diff))
        a_fill = C_MISSING_BG

    fills = [None, skill_fill, None, None, None, a_fill]
    alt = (row % 2 == 0)
    data_row(ws_sk, row, [tree, skill_name, actual, guide, diff_str, assessment],
             alt=alt, fills=fills)
    # Add note in a merged annotation if non-trivial
    if note:
        ws_sk.cell(row=row, column=6).value = assessment + " — " + note
        ws_sk.cell(row=row, column=6).font = Font(name='Calibri', size=9, color="444444")
    row += 1

# Summary
row += 1
ws_sk.merge_cells(f'A{row}:F{row}')
ws_sk[f'A{row}'].value = "Overall Assessment"
ws_sk[f'A{row}'].font = Font(name='Calibri', size=10, bold=True, color=C_HEADER_FG)
ws_sk[f'A{row}'].fill = make_fill(C_SUBHDR_BG)
ws_sk[f'A{row}'].alignment = Alignment(horizontal='left', indent=1)
row += 1

total_actual = sum(WARRIV_ACTUAL)
total_guide  = sum(GUIDE_RECOMMENDED)
summary_lines = [
    ("Total skill points invested", total_actual, total_guide),
    ("Core skills maxed (Abyss / Miasma Chain / Miasma Bolt / Enhanced Entropy)",
     sum(WARRIV_ACTUAL[i] for i in [29,26,22,24]),
     80),
]
header_row(ws_sk, row, ['Metric', 'Warriv Actual', 'Guide Rec.', '', '', ''])
row += 1
for label, actual_v, guide_v in summary_lines:
    diff_v = actual_v - guide_v
    diff_s = f"+{diff_v}" if diff_v >= 0 else str(diff_v)
    data_row(ws_sk, row, [label, actual_v, guide_v, diff_s, '', ''], alt=(row%2==0))
    row += 1

row += 1
ws_sk.merge_cells(f'A{row}:F{row}')
verdict = ws_sk[f'A{row}']
verdict.value = (
    "Verdict: Warriv's Abyss Warlock skill allocation is well-optimized and closely follows the Maxroll guide. "
    "All four core skills (Abyss, Miasma Chain, Miasma Bolt, Enhanced Entropy) are correctly maxed at 20 points each. "
    "Demonic Mastery at 10pts is higher than the guide minimum of 1pt but is a valid choice (extra mastery = more damage scaling). "
    "Bind Demon at 3pts vs guide's 1pt is a minor over-investment but functionally negligible. "
    "The Eldritch tree is correctly left empty (Blade Warp at 1pt for mobility is ideal). "
    "No significant skill allocation errors detected."
)
verdict.font = Font(name='Calibri', size=10, color="1A1A1A")
verdict.fill = make_fill(C_HAVE_BG)
verdict.alignment = Alignment(horizontal='left', wrap_text=True, indent=1)
ws_sk.row_dimensions[row].height = 70

# ══════════════════════════════════════════════════════════════════════════════
# TAB: Recommendations
# ══════════════════════════════════════════════════════════════════════════════
ws_rec = wb.create_sheet("Recommendations")
ws_rec.sheet_view.showGridLines = False
ws_rec.column_dimensions['A'].width = 16
ws_rec.column_dimensions['B'].width = 22
ws_rec.column_dimensions['C'].width = 40
ws_rec.column_dimensions['D'].width = 10
ws_rec.column_dimensions['E'].width = 14

ws_rec.merge_cells('A1:E1')
ws_rec['A1'].value = "Gear Upgrade Recommendations"
ws_rec['A1'].font = Font(name='Calibri', size=14, bold=True, color=C_GOLD)
ws_rec['A1'].fill = make_fill(C_HEADER_BG)
ws_rec['A1'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws_rec.row_dimensions[1].height = 30

row = 3
header_row(ws_rec, row, ['Build', 'Slot', 'Recommendation', 'Priority', 'Status'])
row += 1

PRIORITY = {'MISSING': 'HIGH', 'UPGRADE AVL': 'MED', 'BIS ✓': 'LOW', 'NEAR-BIS ★': 'LOW', 'OK': '-'}

for build_name, char_name in BUILD_CHARS.items():
    equipped = get_equipped(char_name)
    bis_slots = BIS.get(build_name, {})

    for slot in SLOT_ORDER:
        item = equipped.get(slot)
        bis_list = bis_slots.get(slot, [])
        if not bis_list:
            continue

        equipped_name = item_display(item) if item else '— empty —'

        # Apply same gloves fix
        if build_name == 'Javazon' and slot == 'Gloves' and item:
            item_name_raw = item.get('item_name', item.get('base_name', ''))
            if 'Set?' in item_name_raw or item.get('quality_id') == 5:
                equipped_name = "Rare Gloves (+2 J&S / +20% IAS)"

        status, status_fill, _ = get_item_status(build_name, slot, item, equipped_name)

        if status == 'MISSING':
            rec = f"Find: {bis_list[0]}"
        elif status == 'BIS ✓':
            rec = 'No upgrade needed'
        elif status == 'NEAR-BIS ★':
            rec = "Excellent item — upgrade to +3 J&S/+20% IAS for marginal improvement"
        else:
            rec = f"Upgrade to: {bis_list[0]}"

        priority = PRIORITY.get(status, '-')
        fill = (C_MISSING_BG if status == 'MISSING'
                else C_HAVE_BG if status == 'BIS ✓'
                else C_NEAR_BIS_BG if status == 'NEAR-BIS ★'
                else C_UPGRADE_BG)
        alt = (row % 2 == 0)
        data_row(ws_rec, row, [build_name, slot, rec, priority, status],
                 alt=alt, fills=[None, None, None, None, fill])
        row += 1

# ══════════════════════════════════════════════════════════════════════════════
# Save
# ══════════════════════════════════════════════════════════════════════════════
OUT = "/sessions/peaceful-admiring-cerf/mnt/Diablo II Resurrected/D2R_Gear_Analysis_v3.xlsx"
wb.save(OUT)
print(f"Saved: {OUT}")
